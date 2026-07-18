"""
Importador de resultados desde el Excel que exporta Chess-Results
(https://chess-results.com), formato "Cuadro cruzado por clasificación final".

Estructura típica del archivo (ver captura de referencia):
    Fila 1: "De la Base de Datos de Torneos de Chess-Results https://chess-results.com"
    Fila 2: nombre del torneo
    Fila 3: "Última actualización ..."
    Fila 4: "Cuadro cruzado por clasificación final después de N rondas"
    Fila 5: encabezados -> Rk. | Nombre | Elo | FED | 1.Rd | 2.Rd | ... | N.Rd | Pts. | Des 1 | Des 2 | Des 3
    Fila 6+: una fila por jugador.

Cada celda de ronda tiene el formato [numero_rival][color][resultado], p.ej.:
    "29b1"  -> rival de referencia 29, jugó con negras, ganó (1 punto)
    "10w½"  -> rival de referencia 10, jugó con blancas, empató (½ punto)
    "4w0"   -> rival de referencia 4, jugó con blancas, perdió (0 puntos)
    "-½"    -> no tuvo rival esa ronda (bye), sumó ½ punto
    "-"     -> no tuvo rival esa ronda, sin puntos

El "número de referencia" normalmente es el número de orden inicial del
sorteo (SNo), que Chess-Results a veces exporta en una columna aparte y a
veces omite (quedando el propio Rk. como referencia, válido cuando el
archivo no fue reordenado). Este parser detecta automáticamente cuál de
las dos columnas usar.
"""
import re
import unicodedata
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


HEADER_RANK = {"rk.", "rk", "rank"}
HEADER_STARTNO = {"sno", "no.", "no", "stno", "nro", "nro.", "startno", "no.ini.", "no.ini", "noini"}
HEADER_NOMBRE = {"nombre", "name"}
HEADER_ELO = {"elo", "rtg", "rating"}
HEADER_FED = {"fed", "federacion", "federación"}
HEADER_PTS = {"pts.", "pts", "puntos"}
HEADER_CLUB = {"club/ciudad", "club/city", "club", "ciudad", "city", "team", "equipo"}

# Puntos de liga estilo Fórmula 1, según la posición final del jugador en el
# torneo. Del 11° puesto en adelante no suma puntos de liga.
PUNTOS_LIGA_POR_POSICION = {1: 25, 2: 18, 3: 15, 4: 12, 5: 10, 6: 8, 7: 6, 8: 4, 9: 2, 10: 1}

CELL_PAIRING_RE = re.compile(r"^(\d+)\s*([wb])\s*(1|0|½|1/2|0\.5|\+|-)$", re.IGNORECASE)
# "-" o "-½" etc.: bye pedido por el jugador. Un "0" o "1" sueltos (sin rival)
# aparecen quando Chess-Results marca una ronda sin partida jugada (ausencia,
# walkover) sin usar el signo "-": los tratamos igual que un bye.
CELL_BYE_RE = re.compile(r"^-?\s*(1|0|½|1/2|0\.5)?$")
ROUND_HEADER_RE = re.compile(r"^(\d+)\.?\s*rd\.?$")


def _normalizar(texto) -> str:
    """minúsculas, sin tildes, espacios colapsados — para comparar headers/nombres."""
    if texto is None:
        return ""
    txt = str(texto).strip().lower()
    txt = "".join(c for c in unicodedata.normalize("NFKD", txt) if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", txt).strip()


def _score_a_float(raw: str) -> float:
    raw = raw.strip()
    if raw in ("½", "1/2", "0.5"):
        return 0.5
    return float(raw)


class FilaJugadorExcel:
    def __init__(self, ref: str, nombre_crudo: str, elo: Optional[int], fed: Optional[str],
                 pts: Optional[float], rank: Optional[int]):
        self.ref = ref
        self.nombre_crudo = nombre_crudo
        self.elo = elo
        self.fed = fed
        self.pts = pts
        self.rank = rank


class RondaCelda:
    def __init__(self, opp_ref: Optional[str], color: Optional[str], score: Optional[float], bye: bool):
        self.opp_ref = opp_ref
        self.color = color
        self.score = score
        self.bye = bye


class ResultadoParseoExcel:
    def __init__(self):
        self.nombre_torneo: Optional[str] = None
        self.rondas_detectadas: int = 0
        self.jugadores: List[FilaJugadorExcel] = []
        self.partidas_por_ronda: Dict[int, List[Tuple[str, RondaCelda]]] = {}
        self.avisos: List[str] = []


def _encontrar_fila_headers(sheet) -> Tuple[int, Dict[str, int]]:
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        valores = [_normalizar(c) for c in row]
        if any(v in HEADER_RANK for v in valores) and any(v in HEADER_NOMBRE for v in valores):
            columnas = {v: i for i, v in enumerate(valores) if v}
            return row_idx, columnas
    raise ValueError(
        "No se encontró la fila de encabezados (se esperaba una columna 'Rk.' y otra "
        "'Nombre'). ¿Es un archivo exportado por Chess-Results en formato 'Cuadro "
        "cruzado por clasificación final'?"
    )


def parsear_excel_chess_results(file_bytes: bytes) -> ResultadoParseoExcel:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    resultado = ResultadoParseoExcel()

    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
        for cell in row:
            if not isinstance(cell, str):
                continue
            texto = cell.strip()
            bajo = texto.lower()
            if not texto:
                continue
            if "chess-results" in bajo or "actualizacion" in bajo or "actualización" in bajo or "cuadro cruzado" in bajo:
                continue
            resultado.nombre_torneo = texto
            break
        if resultado.nombre_torneo:
            break

    header_row_idx, columnas = _encontrar_fila_headers(sheet)

    idx_rank = next((columnas[h] for h in HEADER_RANK if h in columnas), None)
    idx_startno = next((columnas[h] for h in HEADER_STARTNO if h in columnas), None)
    idx_nombre = next((columnas[h] for h in HEADER_NOMBRE if h in columnas), None)
    idx_elo = next((columnas[h] for h in HEADER_ELO if h in columnas), None)
    idx_fed = next((columnas[h] for h in HEADER_FED if h in columnas), None)
    idx_pts = next((columnas[h] for h in HEADER_PTS if h in columnas), None)

    if idx_nombre is None:
        raise ValueError("No se pudo identificar la columna 'Nombre' en el archivo.")

    ronda_columnas: List[Tuple[int, int]] = []
    for header_norm, idx in columnas.items():
        m = ROUND_HEADER_RE.match(header_norm)
        if m:
            ronda_columnas.append((int(m.group(1)), idx))
    ronda_columnas.sort(key=lambda t: t[0])
    resultado.rondas_detectadas = len(ronda_columnas)

    # Preferimos una columna de número inicial (SNo) si existe; si no, usamos
    # el propio Rk. (válido si el cuadro no fue reordenado tras la última ronda).
    idx_ref = idx_startno if idx_startno is not None else idx_rank
    if idx_ref is None:
        raise ValueError(
            "El archivo no tiene columna 'Rk.' ni de número inicial: no se puede "
            "resolver a qué jugador corresponde cada rival en las rondas."
        )
    if idx_startno is None:
        resultado.avisos.append(
            "El archivo no trae una columna de 'número inicial' (SNo) separada del "
            "ranking final, así que se usó el Rk. como referencia de emparejamiento. "
            "Si el cuadro fue reordenado por clasificación, algunos cruces de ronda "
            "podrían quedar mal asignados: revisá las rondas jugadas del torneo luego de importar."
        )

    filas = list(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True))

    for row in filas:
        if idx_nombre >= len(row):
            continue
        nombre_valor = row[idx_nombre]
        if not nombre_valor or not str(nombre_valor).strip():
            continue

        ref_valor = row[idx_ref] if idx_ref < len(row) else None
        # El Excel de Chess-Results suele traer, después de la última fila de
        # jugadores, filas de pie de página (leyenda, link al torneo, etc.)
        # que repiten el mismo texto en todas las columnas. Esas filas no
        # tienen un número de referencia real, así que las descartamos acá
        # en vez de tratarlas como si fueran un jugador más.
        if isinstance(ref_valor, (int, float)):
            ref = str(int(ref_valor))
        elif isinstance(ref_valor, str) and ref_valor.strip().isdigit():
            ref = ref_valor.strip()
        else:
            continue

        rank_valor = row[idx_rank] if idx_rank is not None and idx_rank < len(row) else None
        elo_valor = row[idx_elo] if idx_elo is not None and idx_elo < len(row) else None
        fed_valor = row[idx_fed] if idx_fed is not None and idx_fed < len(row) else None
        pts_valor = row[idx_pts] if idx_pts is not None and idx_pts < len(row) else None

        pts = None
        if pts_valor not in (None, ""):
            try:
                pts = _score_a_float(str(pts_valor))
            except ValueError:
                resultado.avisos.append(
                    f"No se pudo interpretar el puntaje '{pts_valor}' del jugador '{str(nombre_valor).strip()}'; se dejó sin puntaje."
                )

        jugador = FilaJugadorExcel(
            ref=ref,
            nombre_crudo=str(nombre_valor).strip(),
            elo=int(elo_valor) if isinstance(elo_valor, (int, float)) else None,
            fed=str(fed_valor).strip() if fed_valor else None,
            pts=pts,
            rank=int(rank_valor) if isinstance(rank_valor, (int, float)) else None,
        )
        resultado.jugadores.append(jugador)

        for numero_ronda, idx_col in ronda_columnas:
            if idx_col >= len(row):
                continue
            crudo = row[idx_col]
            if crudo is None or str(crudo).strip() == "":
                continue
            texto = str(crudo).strip()

            m = CELL_PAIRING_RE.match(texto)
            if m:
                opp_ref, color, score_raw = m.groups()
                # "+" y "-" son victoria/derrota por walkover (rival ausente):
                # Chess-Results los usa en vez de "1"/"0" pero valen lo mismo.
                if score_raw == "+":
                    score = 1.0
                elif score_raw == "-":
                    score = 0.0
                else:
                    score = _score_a_float(score_raw)
                celda = RondaCelda(opp_ref=opp_ref, color=color.lower(), score=score, bye=False)
            else:
                m_bye = CELL_BYE_RE.match(texto)
                if m_bye:
                    score_raw = m_bye.group(1)
                    celda = RondaCelda(
                        opp_ref=None, color=None,
                        score=_score_a_float(score_raw) if score_raw else None,
                        bye=True,
                    )
                else:
                    resultado.avisos.append(
                        f"No se pudo interpretar la celda '{texto}' "
                        f"(jugador '{jugador.nombre_crudo}', ronda {numero_ronda})."
                    )
                    continue

            resultado.partidas_por_ronda.setdefault(numero_ronda, []).append((ref, celda))

    return resultado


class FilaClasificacionExcel:
    """Una fila del Excel 'Clasificación Final' (sin rondas, pero con club/ciudad)."""
    def __init__(self, nombre_crudo: str, elo: Optional[int], club: Optional[str]):
        self.nombre_crudo = nombre_crudo
        self.elo = elo
        self.club = club


class ResultadoParseoClasificacion:
    def __init__(self):
        self.filas: List[FilaClasificacionExcel] = []
        self.tiene_columna_club: bool = False
        self.avisos: List[str] = []


def parsear_excel_clasificacion(file_bytes: bytes) -> ResultadoParseoClasificacion:
    """
    Parsea el Excel 'Clasificación Final' de Chess-Results (el que trae el
    listado de jugadores con su columna 'Club/Ciudad', pero sin las rondas
    jugadas). Sirve para completar automáticamente el club de cada jugador,
    en vez de asignarlo a mano uno por uno desde el panel.
    """
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    resultado = ResultadoParseoClasificacion()

    header_row_idx, columnas = _encontrar_fila_headers(sheet)

    idx_nombre = next((columnas[h] for h in HEADER_NOMBRE if h in columnas), None)
    idx_elo = next((columnas[h] for h in HEADER_ELO if h in columnas), None)
    idx_club = next((columnas[h] for h in HEADER_CLUB if h in columnas), None)

    if idx_nombre is None:
        raise ValueError("No se pudo identificar la columna 'Nombre' en el archivo.")
    resultado.tiene_columna_club = idx_club is not None
    if idx_club is None:
        resultado.avisos.append(
            "El archivo no tiene ninguna columna de club/ciudad, así que no se pudo "
            "asignar el club de ningún jugador. ¿Es el archivo 'Clasificación Final' "
            "exportado por Chess-Results?"
        )

    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if idx_nombre >= len(row):
            continue
        nombre_valor = row[idx_nombre]
        if not nombre_valor or not str(nombre_valor).strip():
            continue

        elo_valor = row[idx_elo] if idx_elo is not None and idx_elo < len(row) else None
        club_valor = row[idx_club] if idx_club is not None and idx_club < len(row) else None
        club_txt = str(club_valor).strip() if club_valor and str(club_valor).strip() else None

        resultado.filas.append(
            FilaClasificacionExcel(
                nombre_crudo=str(nombre_valor).strip(),
                elo=int(elo_valor) if isinstance(elo_valor, (int, float)) else None,
                club=club_txt,
            )
        )

    return resultado


def dividir_nombre(nombre_crudo: str) -> Tuple[str, str]:
    """
    Chess-Results exporta el nombre completo como un solo texto (apellido(s)
    primero). No hay forma 100% confiable de separar apellido/nombre sin
    ambigüedad, así que usamos la convención más común: la última palabra es
    el nombre de pila, el resto es apellido. El administrador puede corregirlo
    después desde el panel si hace falta.
    """
    partes = nombre_crudo.split()
    if len(partes) <= 1:
        return nombre_crudo, ""
    return " ".join(partes[:-1]), partes[-1]


def tokens_normalizados(*partes: str) -> frozenset:
    """
    Compara jugadores por el conjunto de palabras de su nombre completo, sin
    importar el orden ("Apellido Nombre" vs "Nombre Apellido") ni la
    puntuación: algunos exports de Chess-Results separan apellido y nombre
    con una coma ("Ferragud, Leandro") y otros no ("Ferragud Leandro"), así
    que sacamos comas/puntos antes de tokenizar para que ambos formatos
    matcheen al mismo jugador.
    """
    texto = _normalizar(" ".join(p for p in partes if p))
    texto = re.sub(r"[,.]", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return frozenset(t for t in texto.split(" ") if t)


def aplicar_importacion(db, torneo, resultado: ResultadoParseoExcel) -> dict:
    """
    Toma el resultado ya parseado del Excel y lo escribe en la base de datos:
      - Matchea (o crea) un Jugador por cada fila, comparando el conjunto de
        palabras del nombre (así no importa si el excel trae "Apellido Nombre"
        y la base tiene nombre/apellido separados en otro orden).
      - Crea/actualiza el ResultadoTorneo (posición final y puntaje obtenido).
      - Crea las Partida de cada ronda (evitando duplicar cada cruce).
    No recalcula ELO ni asigna club: eso se completa a mano desde el panel.
    """
    # Import diferido para evitar dependencia circular con app.models
    from app.models import Jugador, ResultadoTorneo, Partida

    avisos = list(resultado.avisos)

    # Índice de jugadores existentes por conjunto de tokens del nombre completo.
    existentes = db.query(Jugador).all()
    indice_existentes: Dict[frozenset, Jugador] = {}
    for j in existentes:
        indice_existentes[tokens_normalizados(j.nombre, j.apellido)] = j

    ref_a_jugador: Dict[str, Jugador] = {}
    jugadores_creados = 0
    jugadores_encontrados = 0
    siguiente_id_num = 1

    def _generar_id_lma() -> str:
        nonlocal siguiente_id_num
        while True:
            candidato = f"LMA-{siguiente_id_num:05d}"
            siguiente_id_num += 1
            if not db.query(Jugador).filter(Jugador.id_lma == candidato).first():
                return candidato

    for fila in resultado.jugadores:
        clave = tokens_normalizados(fila.nombre_crudo)
        jugador = indice_existentes.get(clave)

        if jugador is None:
            nombre, apellido = dividir_nombre(fila.nombre_crudo)
            id_lma = _generar_id_lma()
            # Si el Excel trae un Elo de FIDE/rating previo lo usamos como
            # punto de partida; si no, arranca en el default de un jugador
            # nuevo (ver app/core/elo.py).
            from app.core.elo import ELO_INICIAL
            elo_inicial = fila.elo or ELO_INICIAL
            jugador = Jugador(
                id_lma=id_lma,
                nombre=nombre or fila.nombre_crudo,
                apellido=apellido or "",
                elo_blitz=elo_inicial,
                elo_rapida=elo_inicial,
                elo_clasica=elo_inicial,
                estado="Activo",
            )
            db.add(jugador)
            db.flush()
            indice_existentes[clave] = jugador
            jugadores_creados += 1
            avisos.append(
                f"Se creó un jugador nuevo para '{fila.nombre_crudo}' (id {id_lma}). "
                "Revisá el club y los datos personales desde el panel."
            )
        else:
            jugadores_encontrados += 1

        ref_a_jugador[fila.ref] = jugador

    # ResultadoTorneo: posición final, puntaje del torneo, y puntos de liga.
    # Los puntos de liga se asignan estilo Fórmula 1 según la posición final
    # dentro de ESTE torneo (no el puntaje crudo de partidas ganadas): el 1°
    # se lleva 25, 2° 18, 3° 15, 4° 12, 5° 10, 6° 8, 7° 6, 8° 4, 9° 2, 10° 1,
    # del 11° en adelante no suma. Estos son los puntos que después se suman
    # en la clasificación de la liga, tanto para jugadores como para clubes
    # (el club se lleva la suma de los puntos de sus jugadores).
    resultados_creados = 0
    for fila in resultado.jugadores:
        jugador = ref_a_jugador[fila.ref]
        puntos_liga_f1 = PUNTOS_LIGA_POR_POSICION.get(fila.rank, 0)
        existente = (
            db.query(ResultadoTorneo)
            .filter(
                ResultadoTorneo.id_torneo == torneo.id,
                ResultadoTorneo.id_jugador == jugador.id_lma,
            )
            .first()
        )
        puntos = fila.pts if fila.pts is not None else 0
        if existente:
            existente.posicion_final = fila.rank
            existente.puntuacion_obtenida = puntos
            existente.puntos_liga = puntos_liga_f1
            existente.id_club = jugador.id_club
        else:
            db.add(
                ResultadoTorneo(
                    id_torneo=torneo.id,
                    id_jugador=jugador.id_lma,
                    id_club=jugador.id_club,
                    posicion_final=fila.rank,
                    puntos_liga=puntos_liga_f1,
                    puntuacion_obtenida=puntos,
                    variacion_elo=None,
                )
            )
            resultados_creados += 1

    # Partidas por ronda: cada cruce aparece dos veces (una en cada fila), así
    # que lo deduplicamos con un set de pares ya procesados.
    partidas_creadas = 0
    for numero_ronda, celdas in resultado.partidas_por_ronda.items():
        procesados = set()
        celdas_por_ref = {ref: celda for ref, celda in celdas}

        for ref, celda in celdas:
            jugador = ref_a_jugador.get(ref)
            if jugador is None:
                continue

            if celda.bye:
                clave_par = (ref, None, numero_ronda)
                if clave_par in procesados:
                    continue
                procesados.add(clave_par)
                puntos_txt = celda.score if celda.score is not None else 0
                db.add(
                    Partida(
                        id_torneo=torneo.id,
                        ronda=numero_ronda,
                        id_jugador_blancas=jugador.id_lma,
                        id_jugador_negras=None,
                        resultado=f"bye-{puntos_txt}",
                    )
                )
                partidas_creadas += 1
                continue

            par = tuple(sorted([ref, celda.opp_ref]))
            clave_par = (par[0], par[1], numero_ronda)
            if clave_par in procesados:
                continue
            procesados.add(clave_par)

            rival = ref_a_jugador.get(celda.opp_ref)
            if rival is None:
                avisos.append(
                    f"No se encontró el rival de referencia '{celda.opp_ref}' en la ronda "
                    f"{numero_ronda} (jugador '{jugador.nombre} {jugador.apellido}')."
                )
                continue

            if celda.color == "w":
                blancas, negras = jugador, rival
                score_blancas = celda.score
            else:
                blancas, negras = rival, jugador
                score_blancas = 1 - celda.score if celda.score is not None else None

            if score_blancas is None:
                resultado_txt = "½-½"
            elif score_blancas == 1:
                resultado_txt = "1-0"
            elif score_blancas == 0:
                resultado_txt = "0-1"
            else:
                resultado_txt = "½-½"

            existente_partida = (
                db.query(Partida)
                .filter(
                    Partida.id_torneo == torneo.id,
                    Partida.ronda == numero_ronda,
                    Partida.id_jugador_blancas == blancas.id_lma,
                    Partida.id_jugador_negras == negras.id_lma,
                )
                .first()
            )
            if existente_partida:
                existente_partida.resultado = resultado_txt
            else:
                db.add(
                    Partida(
                        id_torneo=torneo.id,
                        ronda=numero_ronda,
                        id_jugador_blancas=blancas.id_lma,
                        id_jugador_negras=negras.id_lma,
                        resultado=resultado_txt,
                    )
                )
                partidas_creadas += 1

    db.commit()

    return {
        "ok": True,
        "jugadores_creados": jugadores_creados,
        "jugadores_encontrados": jugadores_encontrados,
        "partidas_creadas": partidas_creadas,
        "resultados_creados": resultados_creados,
        "rondas_detectadas": resultado.rondas_detectadas,
        "avisos": avisos,
    }


def aplicar_clasificacion_clubes(db, resultado: ResultadoParseoClasificacion) -> dict:
    """
    Toma el resultado ya parseado del Excel 'Clasificación Final' (el que
    trae la columna 'Club/Ciudad') y para cada fila:
      - matchea al jugador por nombre igual que en aplicar_importacion (o lo
        crea si todavía no existe, por si este archivo se sube antes que el
        Cuadro Cruzado),
      - busca el club por nombre (sin importar mayúsculas/espacios) o lo crea
        si es la primera vez que aparece,
      - le asigna ese club al jugador.
    No toca resultados, ELO ni partidas: de eso se encarga el Cuadro Cruzado.
    """
    from app.models import Jugador, Club

    avisos = list(resultado.avisos)

    existentes = db.query(Jugador).all()
    indice_existentes: Dict[frozenset, Jugador] = {}
    for j in existentes:
        indice_existentes[tokens_normalizados(j.nombre, j.apellido)] = j

    clubes_existentes = db.query(Club).all()
    indice_clubes: Dict[str, Club] = {_normalizar(c.nombre): c for c in clubes_existentes}

    siguiente_id_num = 1

    def _generar_id_lma() -> str:
        nonlocal siguiente_id_num
        while True:
            candidato = f"LMA-{siguiente_id_num:05d}"
            siguiente_id_num += 1
            if not db.query(Jugador).filter(Jugador.id_lma == candidato).first():
                return candidato

    jugadores_creados = 0
    jugadores_actualizados = 0
    clubes_creados = 0

    for fila in resultado.filas:
        clave = tokens_normalizados(fila.nombre_crudo)
        jugador = indice_existentes.get(clave)

        if jugador is None:
            nombre, apellido = dividir_nombre(fila.nombre_crudo)
            id_lma = _generar_id_lma()
            from app.core.elo import ELO_INICIAL
            elo_inicial = fila.elo or ELO_INICIAL
            jugador = Jugador(
                id_lma=id_lma,
                nombre=nombre or fila.nombre_crudo,
                apellido=apellido or "",
                elo_blitz=elo_inicial,
                elo_rapida=elo_inicial,
                elo_clasica=elo_inicial,
                estado="Activo",
            )
            db.add(jugador)
            db.flush()
            indice_existentes[clave] = jugador
            jugadores_creados += 1

        if not fila.club:
            continue

        clave_club = _normalizar(fila.club)
        club = indice_clubes.get(clave_club)
        if club is None:
            club = Club(nombre=fila.club)
            db.add(club)
            db.flush()
            indice_clubes[clave_club] = club
            clubes_creados += 1

        if jugador.id_club != club.id:
            jugador.id_club = club.id
            jugadores_actualizados += 1

    db.commit()

    return {
        "ok": True,
        "jugadores_creados": jugadores_creados,
        "jugadores_actualizados": jugadores_actualizados,
        "clubes_creados": clubes_creados,
        "avisos": avisos,
    }
